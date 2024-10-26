from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd
import time
from openpyxl import load_workbook


driver = webdriver.Chrome()

# Array to store all extracted texts
extracted_texts = []

# List of nairaland URLs to scrape
urls = [
    "https://www.nairaland.com/7920791/drove-lagos-akure-2450-using",
    "https://www.nairaland.com/7920791/drove-lagos-akure-2450-using/4",
    "https://www.nairaland.com/7920791/drove-lagos-akure-2450-using/2",
    "https://www.nairaland.com/8247057/buy-petrol-1000-per-litre/",
    "https://www.nairaland.com/7950515/cng-car-conversion-really-worth",
    "https://www.nairaland.com/7920791/drove-lagos-akure-2450-using/5",
    "https://www.nairaland.com/7920791/drove-lagos-akure-2450-using/3",
    "https://www.nairaland.com/7920791/drove-lagos-akure-2450-using/1",
    "https://www.nairaland.com/8249174/cng-why-theres-long-queue/",
    "https://www.nairaland.com/8036511/cng-domestic-use",
    "https://www.nairaland.com/7920791/drove-lagos-akure-2450-using/7",
    "https://www.nairaland.com/8247057/buy-petrol-1000-per-litre/1",
    "https://www.nairaland.com/7950515/cng-car-conversion-really-worth/1",
    "https://www.nairaland.com/8036511/cng-domestic-use/1",
    "https://www.nairaland.com/search/cng/0/0/0/1",
    "https://www.nairaland.com/8249174/cng-why-theres-long-queue/1",
    "https://www.nairaland.com/8247057/buy-petrol-1000-per-litre/2",
    "https://www.nairaland.com/8036511/cng-domestic-use/2",
    "https://www.nairaland.com/7950515/cng-car-conversion-really-worth/2",
    "https://www.nairaland.com/8241215/cng-scam-see-long-queue#132472196",
    "https://www.nairaland.com/8247057/buy-petrol-1000-per-litre/3",
    "https://www.nairaland.com/search/cng/0/0/0/2",
    "https://www.nairaland.com/8247057/buy-petrol-1000-per-litre/4",
    "https://www.nairaland.com/8247057/buy-petrol-1000-per-litre/5",
    "https://www.nairaland.com/8247057/buy-petrol-1000-per-litre/6",
    "https://www.nairaland.com/search/cng/0/0/0/3",
    "https://www.nairaland.com/8248070/tinubu-wants-filling-stations-converted/1",
    "https://www.nairaland.com/8180942/danfo-buses-lagos-getting-cng/",
    "https://www.nairaland.com/8180942/danfo-buses-lagos-getting-cng/1",
    "https://www.nairaland.com/search/cng/0/0/0/4",
    "https://www.nairaland.com/8180942/danfo-buses-lagos-getting-cng/2"

]

for url in urls:
    try:
        # Load the webpage
        driver.get(url)

        # Wait for the page to fully load
        time.sleep(5)

        # Get the page source
        page_source = driver.page_source

        # Parse the page source using BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        # Find all blockquote tags
        blockquotes = soup.find_all('blockquote')

        # Find all div tags with class "narrow"
        narrow_divs = soup.find_all('div', class_='narrow')

        # Extract text from blockquotes and add them to the array
        for blockquote in blockquotes:
            extracted_texts.append(blockquote.get_text(strip=True))

        # Extract text from divs with class "narrow" and add them to the array
        for div in narrow_divs:
            extracted_texts.append(div.get_text(strip=True))

    except Exception as e:
        print(f"Error loading {url}: {e}")
        # Retry after a delay if an error occurs
        time.sleep(10)
        continue

driver.quit()

df = pd.DataFrame(extracted_texts, columns=["comments"])

file_path = "model_input.xlsx"

try:
    # Load the existing workbook
    book = load_workbook(file_path)


    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        start_row = writer.sheets['Sheet1'].max_row

        df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=start_row)
    
    print(f"Successfully appended data to {file_path}")

except FileNotFoundError:
    # If the file doesn't exist, create a new one
    df.to_excel(file_path, index=False)
    print(f"File not found. Created a new file: {file_path}")

print("Data successfully saved to 'model_input.xlsx'")
