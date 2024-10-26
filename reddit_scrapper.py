import praw
import dotenv
import requests
import pandas as pd
from openpyxl import load_workbook

hf_bearer_token = dotenv.get_key(dotenv.find_dotenv(), "HF_BEARER_TOKEN")
client_id = dotenv.get_key(dotenv.find_dotenv(), "REDDIT_CLIENT_ID")
client_secret = dotenv.get_key(dotenv.find_dotenv(), "REDDIT_CLIENT_SECRET")

# API_URL = "https://api-inference.huggingface.co/models/distilbert/distilbert-base-uncased-finetuned-sst-2-english" 
# API_URL = "https://api-inference.huggingface.co/models/finiteautomata/bertweet-base-sentiment-analysis"
API_URL = "https://api-inference.huggingface.co/models/dipawidia/xlnet-base-cased-product-review-sentiment-analysis"
headers = {"Authorization": hf_bearer_token}

def query(payload):
    response = requests.post(API_URL, headers=headers, json=payload)
    return response.json()

# Reddit API credentials
reddit = praw.Reddit(
    client_id=client_id,
    client_secret=client_secret,
    user_agent='my_reddit_scraper/0.1 by u/cx'
)

comments_list = []

urls = [
    "https://www.reddit.com/r/Columbus/comments/1g4asjm/cota_buses_why_cng/",
    "https://www.reddit.com/r/CarsIndia/comments/y3n0gb/is_it_worth_going_for_cng_petrol_cars_specially/",
    "https://www.reddit.com/r/CarsIndia/comments/18z9gnh/cng_vs_petroldiesel/",
    "https://www.reddit.com/r/CarsIndia/comments/1eokb94/whats_this_crazy_trend_or_getting_after_market/",
]

for url in urls:
    # Specify the URL of the Reddit page with comments
    submission = reddit.submission(url=url)

    # Fetch the top comments
    submission.comments.replace_more(limit=0)
    comments = submission.comments.list()

    for comment in comments:
        comments_list.append(comment.body)

# Create a DataFrame with a single column 'Comments'
df = pd.DataFrame(comments_list, columns=['Comments'])

# Path to the Excel file
file_path = "model_input.xlsx"

try:
    # Try to load the existing workbook and append new data
    book = load_workbook(file_path)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        # Get the last row in the existing file
        start_row = writer.sheets['Sheet1'].max_row

        # Append new data without writing the header
        df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=start_row)
    print(f"Successfully appended new comments to {file_path}")

except FileNotFoundError:
    # If the file doesn't exist, create a new one
    df.to_excel(file_path, index=False)
    print(f"No existing file found. Created a new file: {file_path}")


